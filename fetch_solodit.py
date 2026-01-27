#!/usr/bin/env python3
"""Solodit Findings Crawler - Fetches security audit findings from Solodit API."""

import json
import os
import sys
import time
from datetime import datetime, timezone
from pathlib import Path

import requests
from openpyxl import Workbook, load_workbook

# Configuration
API_URL = "https://solodit.cyfrin.io/api/v1/solodit/findings"
PAGE_SIZE = 100
REQUEST_DELAY = 3  # seconds between requests
MAX_RETRIES = 3
RETRY_DELAY = 5  # initial retry delay in seconds

# File paths
SCRIPT_DIR = Path(__file__).parent
STATE_FILE = SCRIPT_DIR / "state.json"
HIGH_MEDIUM_FILE = SCRIPT_DIR / "high_medium_findings.xlsx"
LOW_GAS_FILE = SCRIPT_DIR / "low_gas_findings.xlsx"

# Column headers for xlsx
HEADERS = [
    "id", "slug", "title", "impact",
    "quality_score", "general_score",
    "firm_name", "protocol_name",
    "content", "summary",
    "tags", "finders", "finders_count",
    "source_link", "github_link", "pdf_link",
    "contest_link", "contest_prize_txt",
    "report_date"
]


def get_api_key():
    """Get API key from environment variable."""
    api_key = os.environ.get("SOLODIT_API_KEY")
    if not api_key:
        print("Error: SOLODIT_API_KEY environment variable not set")
        sys.exit(1)
    return api_key


def load_state():
    """Load state from state.json file."""
    if STATE_FILE.exists():
        with open(STATE_FILE, "r") as f:
            return json.load(f)
    return {
        "high_medium_max_id": "0",
        "low_gas_max_id": "0",
        "last_run": None,
        "high_medium_count": 0,
        "low_gas_count": 0
    }


def save_state(state):
    """Save state to state.json file."""
    state["last_run"] = datetime.now(timezone.utc).isoformat()
    with open(STATE_FILE, "w") as f:
        json.dump(state, f, indent=2)


def extract_tags(finding):
    """Extract tags from finding as comma-separated string."""
    tags = finding.get("issues_issuetagscore", [])
    return ", ".join(t.get("tags_tag", {}).get("title", "") for t in tags if t.get("tags_tag"))


def extract_finders(finding):
    """Extract finders from finding as comma-separated string."""
    finders = finding.get("issues_issue_finders", [])
    return ", ".join(f.get("wardens_warden", {}).get("handle", "") for f in finders if f.get("wardens_warden"))


def finding_to_row(finding):
    """Convert a finding dict to a row list for xlsx."""
    return [
        finding.get("id", ""),
        finding.get("slug", ""),
        finding.get("title", ""),
        finding.get("impact", ""),
        finding.get("quality_score", ""),
        finding.get("general_score", ""),
        finding.get("firm_name", ""),
        finding.get("protocol_name", ""),
        finding.get("content", ""),
        finding.get("summary", ""),
        extract_tags(finding),
        extract_finders(finding),
        finding.get("finders_count", ""),
        finding.get("source_link", ""),
        finding.get("github_link", ""),
        finding.get("pdf_link", ""),
        finding.get("contest_link", ""),
        finding.get("contest_prize_txt", ""),
        str(finding.get("report_date", "")) if finding.get("report_date") else ""
    ]


def fetch_page(api_key, page, impact_filter):
    """Fetch a single page of findings with retry logic."""
    headers = {
        "Content-Type": "application/json",
        "X-Cyfrin-API-Key": api_key
    }
    payload = {
        "page": page,
        "pageSize": PAGE_SIZE,
        "filters": {
            "impact": impact_filter,
            "sortField": "Recency",
            "sortDirection": "Desc"
        }
    }

    for attempt in range(MAX_RETRIES):
        try:
            response = requests.post(API_URL, headers=headers, json=payload, timeout=30)

            if response.status_code == 429:
                # Rate limited - wait and retry
                reset_time = int(response.headers.get("X-RateLimit-Reset", time.time() + 60))
                wait_time = max(reset_time - time.time(), 60)
                print(f"  Rate limited. Waiting {wait_time:.0f}s...")
                time.sleep(wait_time)
                continue

            response.raise_for_status()
            data = response.json()

            # Check rate limit remaining
            rate_limit = data.get("rateLimit", {})
            remaining = rate_limit.get("remaining", 20)
            if remaining < 2:
                reset_time = rate_limit.get("reset", time.time() + 60)
                wait_time = max(reset_time - time.time(), 1)
                print(f"  Rate limit low ({remaining}). Waiting {wait_time:.0f}s...")
                time.sleep(wait_time)

            return data

        except requests.exceptions.RequestException as e:
            if attempt < MAX_RETRIES - 1:
                delay = RETRY_DELAY * (2 ** attempt)
                print(f"  Request failed: {e}. Retrying in {delay}s...")
                time.sleep(delay)
            else:
                raise

    return None


def create_workbook_with_headers():
    """Create a new workbook with headers."""
    wb = Workbook()
    ws = wb.active
    ws.append(HEADERS)
    return wb


def load_or_create_workbook(filepath):
    """Load existing workbook or create new one with headers."""
    if filepath.exists():
        return load_workbook(filepath)
    return create_workbook_with_headers()


def get_existing_ids(filepath):
    """Get set of existing IDs from xlsx file."""
    if not filepath.exists():
        return set()

    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    ids = set()
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0]:
            ids.add(str(row[0]))
    wb.close()
    return ids


def fetch_category(api_key, impact_filter, xlsx_file, max_id_key, state):
    """Fetch all findings for a category (HIGH/MEDIUM or LOW/GAS)."""
    category_name = "HIGH/MEDIUM" if "HIGH" in impact_filter else "LOW/GAS"
    print(f"\n{'='*50}")
    print(f"Fetching {category_name} findings...")
    print(f"{'='*50}")

    current_max_id = int(state.get(max_id_key, "0"))
    is_initial_fetch = current_max_id == 0

    if is_initial_fetch:
        print("Initial fetch - will download all findings")
        # Create fresh workbook
        wb = create_workbook_with_headers()
    else:
        print(f"Incremental fetch - max_id: {current_max_id}")
        # Load existing workbook
        wb = load_or_create_workbook(xlsx_file)

    ws = wb.active

    page = 1
    new_findings = []
    new_max_id = current_max_id
    total_fetched = 0
    should_stop = False

    while not should_stop:
        print(f"  Fetching page {page}...", end=" ", flush=True)

        data = fetch_page(api_key, page, impact_filter)
        if not data:
            print("Failed to fetch page")
            break

        findings = data.get("findings", [])
        metadata = data.get("metadata", {})
        total_results = metadata.get("totalResults", 0)
        total_pages = metadata.get("totalPages", 0)

        if page == 1:
            print(f"Total: {total_results} findings, {total_pages} pages")
        else:
            print(f"Got {len(findings)} findings")

        if not findings:
            break

        for finding in findings:
            finding_id = int(finding.get("id", 0))

            # Update max ID
            if finding_id > new_max_id:
                new_max_id = finding_id

            # For incremental fetch, stop when we hit existing data
            if not is_initial_fetch and finding_id <= current_max_id:
                print(f"  Reached existing data at ID {finding_id}")
                should_stop = True
                break

            new_findings.append(finding)
            total_fetched += 1

        # Check if we've reached the last page
        if page >= total_pages:
            break

        page += 1
        time.sleep(REQUEST_DELAY)

    # Append new findings to worksheet
    if new_findings:
        print(f"\n  Adding {len(new_findings)} new findings to xlsx...")
        for finding in reversed(new_findings):  # Add oldest first so newest is at bottom
            ws.append(finding_to_row(finding))

        wb.save(xlsx_file)
        print(f"  Saved to {xlsx_file}")
    else:
        print("\n  No new findings to add")

    wb.close()

    # Update state
    state[max_id_key] = str(new_max_id)
    count_key = "high_medium_count" if "HIGH" in impact_filter else "low_gas_count"
    state[count_key] = state.get(count_key, 0) + len(new_findings)

    return len(new_findings)


def main():
    """Main entry point."""
    print("="*60)
    print("Solodit Findings Crawler")
    print(f"Started at: {datetime.now(timezone.utc).isoformat()}")
    print("="*60)

    api_key = get_api_key()
    state = load_state()

    if state.get("last_run"):
        print(f"Last run: {state['last_run']}")
    else:
        print("First run - will perform initial full fetch")

    total_new = 0

    # Fetch HIGH/MEDIUM findings
    total_new += fetch_category(
        api_key,
        ["HIGH", "MEDIUM"],
        HIGH_MEDIUM_FILE,
        "high_medium_max_id",
        state
    )

    # Fetch LOW/GAS findings
    total_new += fetch_category(
        api_key,
        ["LOW", "GAS"],
        LOW_GAS_FILE,
        "low_gas_max_id",
        state
    )

    # Save state
    save_state(state)

    print("\n" + "="*60)
    print(f"Completed at: {datetime.now(timezone.utc).isoformat()}")
    print(f"Total new findings: {total_new}")
    print(f"HIGH/MEDIUM count: {state.get('high_medium_count', 0)}")
    print(f"LOW/GAS count: {state.get('low_gas_count', 0)}")
    print("="*60)


if __name__ == "__main__":
    main()
